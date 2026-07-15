from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border,Side

wb=Workbook()
ws=wb.active

ws["A1"] = "名前"
ws["B1"] = "年齢"
ws["C1"] = "都市"
ws["A2"] = "太郎"
ws["B2"] = 25
ws["C2"] = "東京"

#太字・文字色・サイズ
ws["A1"].font = Font(bold=True, color="FF0000", size=14)

#背景色
ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")

#文字を中央揃え
ws["A1"].alignment = Alignment(horizontal="center")

#罫線
side = Side(style="thin")
ws["A1"].border = Border(left=side, right=side, top=side, bottom=side)

#列の幅を変える
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 10

wb.save("styled.xlsx")
print("styled.xlsxを作成しました!")