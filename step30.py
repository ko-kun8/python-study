import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from datetime import datetime

# 売上データを作る
data = {
    "担当者": ["太郎", "花子", "次郎", "四郎", "太郎", "花子"],
    "商品": ["りんご", "バナナ", "りんご", "みかん", "バナナ", "みかん"],
    "売上": [100, 150, 200, 80, 120, 90]
}
df = pd.DataFrame(data)

# 担当者別集計
summary = df.groupby("担当者")["売上"].sum().reset_index()
summary.columns = ["担当者", "売上合計"]

# Excelファイルを作る
wb = Workbook()
ws = wb.active
ws.title = "売上レポート"

# タイトルを書く
today = datetime.now().strftime("%Y/%m/%d")
ws["A1"] = f"{today} 売上レポート"
ws["A1"].font = Font(bold=True, size=14)

# ヘッダーを書く
ws["A3"] = "担当者"
ws["B3"] = "売上合計"
for cell in [ws["A3"], ws["B3"]]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="2196F3")
    cell.alignment = Alignment(horizontal="center")

# データを書く
for i, row in summary.iterrows():
    ws[f"A{i+4}"] = row["担当者"]
    ws[f"B{i+4}"] = row["売上合計"]

# 列幅を設定
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 15

# グラフを作る
data_ref = Reference(ws, min_col=2, min_row=4,
                     max_row=3+len(summary))
cats = Reference(ws, min_col=1, min_row=4,
                 max_row=3+len(summary))
chart = BarChart()
chart.title = "担当者別売上"
chart.add_data(data_ref)
chart.set_categories(cats)
ws.add_chart(chart, "D3")

# 保存
filename = r"C:\Users\ka06_\OneDrive\デスクトップ\python練習\売上レポート_{}.xlsx".format(
    datetime.now().strftime('%Y%m%d')
)
wb.save(filename)
print(f"{filename} を作成しました！")