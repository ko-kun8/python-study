import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import time

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}

base_url = "https://books.toscrape.com/catalogue/page-{}.html"
data = []

for page in range(1, 3):
    url = base_url.format(page)
    print(f"取得中: {url}")

    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    books = soup.select("article.product_pod")
    for book in books:
        title = book.select_one("h3 a")["title"]
        price = book.select_one(".price_color").text.encode("latin1").decode("utf-8")
        data.append({"書籍名": title, "価格": price})

    time.sleep(1)

df = pd.DataFrame(data)

#Excelに保存
df.to_excel("books_styled.xlsx", index=False)

#装飾する
wb = load_workbook("books_styled.xlsx")
ws = wb.active

#ヘッダーを装飾
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center")

#列幅を調整
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 15

wb.save("books_styled.xlsx")
print("books_styled.xlsx を保存しました！")