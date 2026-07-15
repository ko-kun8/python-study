from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "シート１"

ws["A1"] = "名前"
ws["B1"] = "年齢"
ws["C1"] = "都市"

ws["A2"] = "太郎"
ws["B2"] = "25"
ws["C2"] = "東京"

ws["A3"] = "花子"
ws["B3"] = "30"
ws["C3"] = "大阪"

wb.save("test.xlsx")
print("test.xlsx を作りました！")

wb2 = load_workbook("test.xlsx")
ws2 = wb2.active

for row in ws2.iter_rows(values_only=True):
    print(row)