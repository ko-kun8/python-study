from openpyxl import load_workbook

#テンプレートを読み込む
wb = load_workbook("template.xlsx")
ws = wb.active

#データを流し込む
ws["B1"] = "株式会社テスト"
ws["B2"] = "太郎"
ws["B3"] = 300000

#別名で保存
wb.save("output_report.xlsx")
print("output_report.xlsxを作成しました!")