from openpyxl import Workbook

wb = Workbook()

#最初のシートの名前を変える
ws1 = wb.active
ws1.title = "１月"
ws1["A1"] = "１月の売上"
ws1["B1"] = 100

#新しいシートを追加
ws2 = wb.create_sheet("２月")
ws2["A1"] = "２月の売上"
ws2["B1"] = 150

ws3 = wb.create_sheet("３月")
ws3["A1"] = "３月の売上"
ws3["B1"] = 200

# 全シートの名前を表示
print(wb.sheetnames)

#全シートをループして中身を表示
for sheet in wb:
    print(sheet.title, sheet["B1"].value)

#シートをコピーする
ws_copy = wb.copy_worksheet(ws1)
ws_copy.title = "１月コピー"

wb.save("sheets.xlsx")
print("sheets.xlsxを作成しました！")