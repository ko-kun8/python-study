from unicodedata import category
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active

#データを入力
ws["A1"] = "名前"
ws["B1"] = "売上"
ws["A2"] = "太郎"
ws["B2"] = 100
ws["A3"] = "花子"
ws["B3"] = 150
ws["A4"] = "次郎"
ws["B4"] = 80
ws["A5"] = "四郎"
ws["B5"] = 200

#グラフのデータ範囲を指定
data = Reference(ws, min_col=2, min_row=2, max_row=5)
categories = Reference(ws, min_col=1, min_row=2, max_row=5)

#棒グラフを作る
chart = BarChart()
chart.title = "売上グラフ"
chart.y_axis.title = "売上"
chart.x_axis.title = "名前"
chart.add_data(data)
chart.set_categories(categories)

#シートに貼り付ける
ws.add_chart(chart, "D1")

wb.save("chart.xlsx")
print("chart.xlsx　を作成しました！")